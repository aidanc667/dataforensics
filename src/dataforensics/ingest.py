import csv
import datetime
import io
import json
import sys
from pathlib import Path

from charset_normalizer import from_path

# The stdlib csv module defaults to a 128KB per-field cap, meant to protect
# a streaming reader from a pathological input. We already hold the whole
# line in memory before handing it to csv.reader (see split_delimited_line),
# so the cap only serves to crash on legitimate long free-text/notes fields
# (clinical notes, "other, please specify" survey responses) with an
# unhandled _csv.Error instead of a real ingest error message. sys.maxsize
# can overflow the C long the limit is stored in on some platforms, so back
# off until it's accepted.
_field_size_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(_field_size_limit)
        break
    except OverflowError:
        _field_size_limit //= 2

_CANDIDATE_DELIMITERS = [",", "\t", ";", "|"]
_FOOTER_MISMATCH_RUN = 2


class DuplicateHeaderError(Exception):
    """Raised when a CSV header row contains the same column name more than
    once. Every header-parsing site in this codebase (build_data_dictionary,
    read_rows, and cli._read_header_and_row_count) shares this check via
    check_header_has_no_duplicates -- without it, dict-based row/column
    construction (dict(zip(...)), {name: [] for name in header}) silently
    collapses same-named columns, and earlier-occurring columns' data is
    lost with no error and exit 0. This is a malformed-input-file condition,
    not a config problem; callers should map it to a runtime/IO failure
    (exit code 3), not a config error (exit code 2)."""


def check_header_has_no_duplicates(header: list[str]) -> None:
    seen: set[str] = set()
    duplicates: list[str] = []
    for name in header:
        if name in seen and name not in duplicates:
            duplicates.append(name)
        seen.add(name)
    if duplicates:
        names = ", ".join(f"'{d}'" for d in duplicates)
        raise DuplicateHeaderError(
            f"Duplicate column name(s) in header: {names} — every same-named "
            "column after the first would silently lose its data if this were "
            "allowed to proceed"
        )


def deduplicate_header(header: list[str]) -> list[str]:
    """Deterministically rename duplicate header entries by appending a
    positional suffix (second 'name' -> 'name_2', third -> 'name_3', ...).

    This is NOT called anywhere in the default CLI path -- `scan`/`harmonize`
    still refuse via `check_header_has_no_duplicates` by default, since a
    silent rename could surprise a scripted/automated caller. It exists for
    an interactive caller (e.g. a UI) to offer as an explicit, opt-in,
    clearly-logged remediation: renaming is lossless (every column survives
    under a distinguishable name) and fully deterministic, unlike merging or
    dropping, so it's safe to offer as a one-click fix -- but only ever with
    the human choosing it, never automatically.
    """
    seen_counts: dict[str, int] = {}
    result = []
    for name in header:
        seen_counts[name] = seen_counts.get(name, 0) + 1
        if seen_counts[name] == 1:
            result.append(name)
        else:
            result.append(f"{name}_{seen_counts[name]}")
    return result


def detect_encoding(path: Path) -> str:
    result = from_path(str(path)).best()
    if result is None:
        return "utf-8"

    encoding = result.encoding
    # Normalize encoding name spelling to use hyphens instead of underscores
    # (e.g. "utf_8" -> "utf-8"). This only reformats the string charset_normalizer
    # returned; it never changes which encoding is being reported.
    encoding = encoding.replace("_", "-")

    return encoding


def read_source_lines(path: Path) -> tuple[list[str], str]:
    """Read `path` as text and split into lines, dropping any fully blank
    (whitespace-only) line before it reaches delimiter detection or
    strip_footer's field-count heuristic.

    A blank line in a CSV/TSV is formatting noise (e.g. a stray trailing
    newline, or a spacer someone added by hand), not a genuine empty
    record -- this matches how virtually every real-world CSV reader
    treats it (pandas' `skip_blank_lines=True` default included). Left
    unfiltered, a blank line does two bad things: it counts as a
    "0 vs N fields" mismatch toward strip_footer's footer-detection run
    (turning an innocuous blank line into an accidental trigger for
    truncating the rest of the file), and once past that, it becomes a
    spurious all-null row once zip_longest pads it out to the header's
    width in read_rows -- a real observed bug where mid-file blank lines
    silently inflated the row count with fake empty records.

    The single shared primitive every delimited-text call site in this
    codebase should read a file's lines through, instead of each one
    hand-rolling `detect_encoding` + `.read_text().splitlines()` (and, as
    happened before this function existed, each one being independently
    exposed to the blank-line bug above). Returns (data_lines, encoding).
    """
    encoding = detect_encoding(path)
    try:
        # A leading U+FEFF byte-order mark decodes as a literal character
        # under a plain "utf-8"/"utf-16" codec (only the "-sig" codec variants
        # strip it automatically), and detect_encoding can report the plain
        # name even when the file carries a BOM (e.g. Excel's "CSV UTF-8"
        # export always writes one). Left in place it silently prefixes the
        # first header's name, breaking every rules-file column match against
        # that column with no visible error.
        raw_lines = path.read_text(encoding=encoding).lstrip("\ufeff").splitlines()
    except UnicodeDecodeError as exc:
        # detect_encoding's best guess (or its "utf-8" fallback when it
        # couldn't guess at all) still failed to decode every byte -- this
        # is not a CSV/TSV encoding quirk at that point, it's very likely
        # a binary file (an image, a PDF, a genuinely corrupted upload)
        # with a text-file extension. Fail cleanly here with a message the
        # existing IngestFormatError handling already shows the user,
        # instead of letting a raw UnicodeDecodeError traceback surface.
        raise IngestFormatError(
            f"{path.name} doesn't look like a text file (tried decoding as {encoding}: {exc}). "
            "This is likely a binary file (e.g. an image or PDF) with a .csv/.tsv extension, "
            "or a corrupted download -- re-export it as plain-text CSV/TSV and re-upload."
        ) from exc
    return [line for line in raw_lines if line.strip() != ""], encoding


def detect_delimiter(sample_lines: list[str]) -> str:
    non_empty = [line for line in sample_lines if line.strip()]
    if not non_empty:
        return ","

    best_delim = ","
    best_score = -1
    for delim in _CANDIDATE_DELIMITERS:
        counts = [line.count(delim) for line in non_empty]
        if counts[0] == 0:
            continue
        consistent = sum(1 for c in counts if c == counts[0])
        score = consistent * counts[0]
        if score > best_score:
            best_score = score
            best_delim = delim
    return best_delim


def split_delimited_line(line: str, delimiter: str) -> list[str]:
    """Splits one line of delimited text into fields the way a real CSV/TSV
    dialect parser would -- via Python's ``csv`` module -- so a quoted field
    containing the delimiter itself (e.g. ``"Delta Clinic, North"`` in a
    comma-delimited file) counts and splits as ONE field, not two. This is
    the single shared primitive every delimited-text field-count or
    field-split in this codebase should go through, so quoting is handled
    consistently everywhere rather than each call site hand-rolling its own
    ``.split(delimiter)``/``.count(delimiter)``.

    An empty line has no fields under csv.reader (it yields nothing at
    all, unlike ``"".split(delimiter)`` which returns ``['']``) -- returning
    ``[""]`` here instead preserves the old single-empty-field behavior a
    blank line in the middle of a file previously had, rather than raising
    StopIteration on input that used to be handled, just unusually.
    """
    if line == "":
        return [""]
    return next(csv.reader(io.StringIO(line), delimiter=delimiter))


def join_delimited_line(fields: list[str], delimiter: str) -> str:
    """The write-side counterpart to split_delimited_line: joins fields back
    into one delimited line via csv.writer, so a field that itself contains
    the delimiter (or a quote character) gets correctly re-quoted instead of
    corrupting the line a naive ``delimiter.join(fields)`` would produce.
    """
    buffer = io.StringIO()
    csv.writer(buffer, delimiter=delimiter, lineterminator="").writerow(fields)
    return buffer.getvalue()


def strip_footer(lines: list[str], delimiter: str) -> tuple[list[str], list[str]]:
    """Split ``lines`` into (data_lines, stripped_lines) by looking for a run
    of consecutive lines whose real (quote-aware) field count disagrees with
    the header's (e.g. a CDC WONDER-style "Query Parameters:" footer block).

    A candidate run is only treated as a genuine footer if EVERY line from
    its start to end-of-file also mismatches the header's field count. A
    real trailing footer never has ordinary data mixed back in after it
    starts -- if a normal-shaped line reappears anywhere after the run
    (e.g. a stray ragged/short row in the middle of an otherwise clean
    file), this isn't a footer, it's noise in the MIDDLE of the file, and
    nothing is stripped. This was a real observed bug: a single ragged
    row happening to sit right after a mismatched line was enough to
    silently discard every genuine row after it, all the way to
    end-of-file. Downstream code (read_rows' zip_longest padding) already
    handles a ragged row without losing it, so refusing to strip here
    never leaves that row unhandled -- it just stops treating it as the
    start of a footer.
    """
    if not lines:
        return [], []

    header_fields = len(split_delimited_line(lines[0], delimiter))
    mismatch_start = None
    run_length = 0

    for i in range(1, len(lines)):
        fields = len(split_delimited_line(lines[i], delimiter))
        if fields != header_fields:
            run_length += 1
            if run_length >= _FOOTER_MISMATCH_RUN:
                mismatch_start = i - _FOOTER_MISMATCH_RUN + 1
                break
        else:
            run_length = 0

    if mismatch_start is None:
        return lines, []

    genuinely_footer = all(
        len(split_delimited_line(line, delimiter)) != header_fields for line in lines[mismatch_start:]
    )
    if not genuinely_footer:
        return lines, []

    return lines[:mismatch_start], lines[mismatch_start:]


class IngestFormatError(Exception):
    """Raised when a JSON or Excel input file doesn't match the shape
    DataForensics requires to safely treat it as tabular data -- e.g. JSON
    that isn't an array of flat objects, a JSON field whose value is itself
    an object/array, or a multi-sheet Excel workbook with no sheet chosen.
    Like DuplicateHeaderError, this signals a malformed-input-file
    condition, not a config problem -- callers should map it to exit code 3,
    not exit code 2."""


def detect_file_format(path: Path) -> str:
    """Returns "json", "excel", or "delimited" (the existing CSV/TSV path)
    based on the file extension alone -- no content sniffing. Any
    extension other than .json/.xlsx/.xls (including no extension, or an
    unrecognized one like .txt) is treated as delimited text, preserving
    today's behavior for every file this tool already accepts."""
    suffix = path.suffix.lower()
    if suffix == ".json":
        return "json"
    if suffix in (".xlsx", ".xls"):
        return "excel"
    return "delimited"


def _stringify_cell(value) -> str:
    """Converts one JSON/Excel scalar value to the plain-text form the rest
    of the engine expects, matching what the same value would look like if
    it had come from a CSV cell instead. Used by both read_json_rows and
    read_excel_rows so the two formats produce identical text for
    equivalent values.

    bool is checked before int/float because bool is a subclass of int in
    Python -- without this order, True would stringify as "1" instead of
    "true". datetime.datetime is checked before datetime.date for the same
    subclass reason (datetime.datetime IS-A datetime.date).
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value)


def read_json_rows(path: Path) -> list[dict[str, str]]:
    """Reads a JSON file that must be a top-level array of flat objects
    (e.g. [{"age": 34, "sex": "F"}, ...]) into the same list[dict[str, str]]
    shape read_rows() produces for CSV/TSV. Never guesses at any other
    shape (a bare object, NDJSON, a nested array-under-a-key) -- see
    docs/superpowers/specs/2026-08-26-json-excel-ingest-design.md for why.
    """
    try:
        text = path.read_text(encoding="utf-8")
    except (UnicodeDecodeError, OSError) as exc:
        raise IngestFormatError(f"{path.name} could not be read as UTF-8 JSON: {exc}") from exc
    try:
        data = json.loads(text)
    except (json.JSONDecodeError, RecursionError) as exc:
        raise IngestFormatError(f"{path.name} is not valid JSON: {exc}") from exc

    if not isinstance(data, list):
        raise IngestFormatError(
            f"{path.name}: expected a JSON array of objects (e.g. [{{...}}, {{...}}]), "
            f"got a top-level {type(data).__name__}"
        )

    if not data:
        return []

    header: list[str] = []
    seen_keys: set[str] = set()
    for i, element in enumerate(data):
        if not isinstance(element, dict):
            raise IngestFormatError(
                f"{path.name}: element {i} is not a JSON object (got {type(element).__name__}) "
                "-- every array element must be a flat object of column name -> value"
            )
        for key in element:
            if key not in seen_keys:
                seen_keys.add(key)
                header.append(key)

    rows: list[dict[str, str]] = []
    for i, element in enumerate(data):
        row: dict[str, str] = {}
        for key in header:
            value = element.get(key)
            if isinstance(value, (dict, list)):
                raise IngestFormatError(
                    f"{path.name}: column '{key}' at element {i} is a "
                    f"{type(value).__name__}, not a single value -- DataForensics "
                    "doesn't guess how to flatten nested JSON"
                )
            row[key] = _stringify_cell(value)
        rows.append(row)
    return rows


def _format_is_date_only(fmt_str: str) -> bool:
    """A cell's number-format string is the authoritative signal for "was
    this cell formatted as a date, not a date+time" -- "h" (hour) or "s"
    (second) in the format code means time-of-day is part of the display,
    so only a format lacking both is treated as date-only. Shared by both
    the .xlsx and .xls backends so the heuristic isn't duplicated."""
    fmt = fmt_str.lower()
    return "h" not in fmt and "s" not in fmt


def _xlsx_backend(path: Path):
    import openpyxl

    # openpyxl.load_workbook raises whatever exception the underlying zip/
    # XML parsing hits (zipfile.BadZipFile for a misnamed or corrupted
    # .xlsx, among others) -- none of those are IngestFormatError, so
    # without this catch-and-reraise a malformed file would produce an
    # unhandled traceback instead of the clean error every other
    # malformed-input case in this module already surfaces. Catching bare
    # Exception here is deliberate: this is a library boundary where the
    # set of failure modes for a corrupted binary file isn't enumerable.
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise IngestFormatError(f"{path.name} could not be read as an .xlsx file: {exc}") from exc
    sheet_names = wb.sheetnames

    def get_grid(sheet_name: str) -> list[tuple]:
        grid = []
        for row in wb[sheet_name].iter_rows():
            values = []
            for cell in row:
                value = cell.value
                # openpyxl always hands back datetime.datetime for a date-
                # formatted cell, never datetime.date, even though the cell
                # itself carries no time component -- it never round-trips
                # a plain date as a date. Without this, _stringify_cell (which
                # deliberately formats date and datetime differently) would
                # render every Excel date cell as an ISO datetime string
                # ("2024-01-15T00:00:00") instead of a plain date
                # ("2024-01-15"), unlike the equivalent CSV/JSON value.
                if (
                    isinstance(value, datetime.datetime)
                    and cell.is_date
                    and _format_is_date_only(cell.number_format)
                ):
                    value = value.date()
                values.append(value)
            grid.append(tuple(values))
        return grid

    return sheet_names, get_grid


def _xls_cell_value(cell, datemode, book):
    import xlrd

    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        # XL_CELL_EMPTY is a cell with no record at all; XL_CELL_BLANK is a
        # cell that has formatting (e.g. a border or fill) but no content.
        # Both mean "no value" -- without treating XL_CELL_BLANK the same
        # as XL_CELL_EMPTY, read_excel_rows's fully-blank-row skip (which
        # checks `all(v is None for v in raw_row)`) would never fire for a
        # formatted-but-empty trailing row, turning it into a row of empty
        # strings instead of being skipped.
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        value = xlrd.xldate_as_datetime(cell.value, datemode)
        # Mirrors the .xlsx date-only handling above: xlrd always hands back
        # a full datetime for a date cell too, so a date-only cell would
        # otherwise stringify with a spurious "T00:00:00" suffix. Unlike
        # openpyxl, xlrd doesn't expose a per-cell number_format directly --
        # it has to be looked up via the cell's XF (extended format) record,
        # which requires the workbook to be opened with formatting_info=True.
        xf = book.xf_list[cell.xf_index] if cell.xf_index is not None else None
        if xf is not None:
            fmt = book.format_map.get(xf.format_key)
            if fmt is not None and _format_is_date_only(fmt.format_str):
                value = value.date()
        return value
    return cell.value


def _xls_backend(path: Path):
    import xlrd

    # Mirrors the .xlsx backend's catch above: xlrd.open_workbook raises its
    # own error types (xlrd.XLRDError, or occasionally something else) for a
    # misnamed or corrupted .xls file, none of which is IngestFormatError.
    try:
        book = xlrd.open_workbook(str(path), formatting_info=True)
    except Exception as exc:
        raise IngestFormatError(f"{path.name} could not be read as a .xls file: {exc}") from exc
    sheet_names = book.sheet_names()

    def get_grid(sheet_name: str) -> list[list]:
        ws = book.sheet_by_name(sheet_name)
        return [
            [
                _xls_cell_value(ws.cell(r, c), book.datemode, book)
                for c in range(ws.ncols)
            ]
            for r in range(ws.nrows)
        ]

    return sheet_names, get_grid


def _excel_backend(path: Path):
    if path.suffix.lower() == ".xlsx":
        return _xlsx_backend(path)
    return _xls_backend(path)


def list_excel_sheets(path: Path) -> list[str]:
    """Lists the sheet names in an Excel workbook, in workbook order.
    Exists mainly for an interactive caller (e.g. the Streamlit app) that
    needs to show a sheet picker before calling read_excel_rows."""
    sheet_names, _get_grid = _excel_backend(path)
    return sheet_names


def read_excel_table(path: Path, sheet: str | None = None) -> tuple[list[str], list[list[str]]]:
    """Like read_excel_rows, but returns (header, body_rows) directly
    instead of list[dict[str, str]] -- this lets a caller (dictionary.py's
    _load_table) distinguish a header-only sheet (real header, zero data
    rows) from a genuinely empty sheet (no header at all), a distinction
    read_excel_rows's list[dict] return cannot represent: an empty list
    means "nothing" in both cases. Same multi-sheet-ambiguity and
    duplicate-header behavior as read_excel_rows -- this is the same
    logic, split at an earlier point.
    """
    sheet_names, get_grid = _excel_backend(path)

    if not sheet_names:
        # Defensive hardening for an untrusted-file-upload boundary: a
        # workbook that opens successfully but reports zero sheets should
        # never be possible from a real .xlsx/.xls file, but if one shows
        # up here, indexing sheet_names[0] below would raise a bare
        # IndexError instead of the clean IngestFormatError every other
        # malformed-input case in this module surfaces.
        raise IngestFormatError(f"{path.name} contains no sheets")

    if sheet is None:
        if len(sheet_names) > 1:
            raise IngestFormatError(
                f"{path.name} has multiple sheets ({', '.join(sheet_names)}) -- choose one "
                "(scan/harmonize --sheet NAME; crosswalk mode requires a single-sheet source, "
                "since --sheet applies to only one file at a time)"
            )
        chosen = sheet_names[0]
    else:
        if sheet not in sheet_names:
            raise IngestFormatError(
                f"{path.name} has no sheet named '{sheet}' -- available sheets: "
                f"{', '.join(sheet_names)}"
            )
        chosen = sheet

    grid = get_grid(chosen)
    if not grid:
        return [], []

    header = [_stringify_cell(v) for v in grid[0]]
    check_header_has_no_duplicates(header)

    body_rows: list[list[str]] = []
    for raw_row in grid[1:]:
        if all(v is None for v in raw_row):
            continue
        stringified = [
            _stringify_cell(v)
            for v in list(raw_row) + [None] * (len(header) - len(raw_row))
        ]
        body_rows.append(stringified)
    return header, body_rows


def read_excel_rows(path: Path, sheet: str | None = None) -> list[dict[str, str]]:
    """Reads one sheet of an .xlsx or .xls workbook into a
    list[dict[str, str]], same shape read_rows() produces for CSV/TSV.
    See read_excel_table for the (header, body_rows) form this is built
    from, used directly by dictionary.py's _load_table to preserve a
    header-only sheet's schema (something this list[dict] return cannot
    represent -- an empty list means "no header" and "header but zero
    rows" identically).
    """
    header, body_rows = read_excel_table(path, sheet=sheet)
    if not header:
        return []
    return [dict(zip(header, row)) for row in body_rows]
