import datetime
import json

import pytest

from dataforensics.ingest import IngestFormatError, _stringify_cell, detect_file_format, read_json_rows


def test_detect_file_format_csv():
    assert detect_file_format_from_name("sample.csv") == "delimited"


def test_detect_file_format_tsv():
    assert detect_file_format_from_name("sample.tsv") == "delimited"


def test_detect_file_format_unrecognized_extension_is_delimited():
    assert detect_file_format_from_name("sample.txt") == "delimited"


def test_detect_file_format_json():
    assert detect_file_format_from_name("sample.json") == "json"


def test_detect_file_format_xlsx():
    assert detect_file_format_from_name("sample.xlsx") == "excel"


def test_detect_file_format_xls():
    assert detect_file_format_from_name("sample.xls") == "excel"


def test_detect_file_format_case_insensitive():
    assert detect_file_format_from_name("SAMPLE.JSON") == "json"
    assert detect_file_format_from_name("SAMPLE.XLSX") == "excel"


def detect_file_format_from_name(name: str) -> str:
    from pathlib import Path

    return detect_file_format(Path(name))


def test_ingest_format_error_is_an_exception():
    assert issubclass(IngestFormatError, Exception)


def test_stringify_cell_none_is_empty_string():
    assert _stringify_cell(None) == ""


def test_stringify_cell_bool_is_lowercase():
    assert _stringify_cell(True) == "true"
    assert _stringify_cell(False) == "false"


def test_stringify_cell_int():
    assert _stringify_cell(5) == "5"


def test_stringify_cell_float_non_whole():
    assert _stringify_cell(5.5) == "5.5"


def test_stringify_cell_float_whole_number_has_no_trailing_point_zero():
    # A spreadsheet cell holding "34" should stringify as "34", not "34.0" --
    # both openpyxl and xlrd can hand back a whole number as a float
    # depending on cell formatting, and the two backends must agree.
    assert _stringify_cell(34.0) == "34"


def test_stringify_cell_date():
    assert _stringify_cell(datetime.date(2024, 1, 15)) == "2024-01-15"


def test_stringify_cell_datetime():
    assert _stringify_cell(datetime.datetime(2024, 1, 15, 0, 0, 0)) == "2024-01-15T00:00:00"


def test_stringify_cell_string_passthrough():
    assert _stringify_cell("hello") == "hello"


def test_stringify_cell_bool_checked_before_int():
    # bool is a subclass of int in Python -- True must not stringify as "1".
    assert _stringify_cell(True) != "1"


def test_read_json_rows_happy_path(tmp_path):
    f = tmp_path / "sample.json"
    f.write_text(json.dumps([
        {"participant_id": "001", "age": 34, "sex": "M"},
        {"participant_id": "002", "age": 29, "sex": "F"},
    ]))
    rows = read_json_rows(f)
    assert rows == [
        {"participant_id": "001", "age": "34", "sex": "M"},
        {"participant_id": "002", "age": "29", "sex": "F"},
    ]


def test_read_json_rows_empty_array_returns_empty_list(tmp_path):
    f = tmp_path / "empty.json"
    f.write_text("[]")
    assert read_json_rows(f) == []


def test_read_json_rows_ragged_keys_fill_missing_with_empty_string(tmp_path):
    f = tmp_path / "ragged.json"
    f.write_text(json.dumps([
        {"a": "1", "b": "2"},
        {"a": "3"},
    ]))
    rows = read_json_rows(f)
    assert rows == [
        {"a": "1", "b": "2"},
        {"a": "3", "b": ""},
    ]


def test_read_json_rows_null_value_becomes_empty_string(tmp_path):
    f = tmp_path / "nulls.json"
    f.write_text(json.dumps([{"a": "1", "b": None}]))
    assert read_json_rows(f) == [{"a": "1", "b": ""}]


def test_read_json_rows_preserves_first_seen_key_order(tmp_path):
    f = tmp_path / "order.json"
    f.write_text(json.dumps([
        {"z": "1", "a": "2"},
        {"b": "3"},
    ]))
    rows = read_json_rows(f)
    assert list(rows[0].keys()) == ["z", "a", "b"]


def test_read_json_rows_invalid_json_raises_ingest_format_error(tmp_path):
    f = tmp_path / "broken.json"
    f.write_text("{not valid json")
    with pytest.raises(IngestFormatError):
        read_json_rows(f)


def test_read_json_rows_non_array_top_level_raises(tmp_path):
    f = tmp_path / "object.json"
    f.write_text(json.dumps({"a": "1"}))
    with pytest.raises(IngestFormatError, match="array"):
        read_json_rows(f)


def test_read_json_rows_non_object_element_raises(tmp_path):
    f = tmp_path / "scalars.json"
    f.write_text(json.dumps(["a", "b"]))
    with pytest.raises(IngestFormatError, match="element 0"):
        read_json_rows(f)


def test_read_json_rows_nested_array_value_raises(tmp_path):
    f = tmp_path / "nested.json"
    f.write_text(json.dumps([{"a": "1", "tags": ["x", "y"]}]))
    with pytest.raises(IngestFormatError, match="'tags'"):
        read_json_rows(f)


def test_read_json_rows_nested_object_value_raises(tmp_path):
    f = tmp_path / "nested_obj.json"
    f.write_text(json.dumps([{"a": "1", "detail": {"x": 1}}]))
    with pytest.raises(IngestFormatError, match="'detail'"):
        read_json_rows(f)


import datetime as dt

import openpyxl
import xlwt

from dataforensics.ingest import list_excel_sheets, read_excel_rows


def _write_xlsx(path, rows, sheet_name="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_read_excel_rows_xlsx_happy_path(tmp_path):
    f = tmp_path / "sample.xlsx"
    _write_xlsx(f, [
        ["participant_id", "age", "sex"],
        ["001", 34, "M"],
        ["002", 29, "F"],
    ])
    rows = read_excel_rows(f)
    assert rows == [
        {"participant_id": "001", "age": "34", "sex": "M"},
        {"participant_id": "002", "age": "29", "sex": "F"},
    ]


def test_read_excel_rows_xlsx_stringifies_dates_and_booleans(tmp_path):
    f = tmp_path / "types.xlsx"
    _write_xlsx(f, [
        ["id", "visit_date", "consented"],
        [1, dt.date(2024, 1, 15), True],
    ])
    rows = read_excel_rows(f)
    assert rows == [{"id": "1", "visit_date": "2024-01-15", "consented": "true"}]


def test_read_excel_rows_xlsx_empty_sheet_returns_empty_list(tmp_path):
    f = tmp_path / "empty.xlsx"
    _write_xlsx(f, [])
    assert read_excel_rows(f) == []


def test_read_excel_rows_xlsx_skips_fully_blank_trailing_rows(tmp_path):
    f = tmp_path / "blank_row.xlsx"
    _write_xlsx(f, [
        ["a", "b"],
        ["1", "2"],
        [None, None],
    ])
    assert read_excel_rows(f) == [{"a": "1", "b": "2"}]


def test_read_excel_rows_xlsx_duplicate_header_raises_duplicate_header_error(tmp_path):
    from dataforensics.ingest import DuplicateHeaderError

    f = tmp_path / "dupe.xlsx"
    _write_xlsx(f, [["a", "a"], ["1", "2"]])
    with pytest.raises(DuplicateHeaderError):
        read_excel_rows(f)


def test_read_excel_rows_xlsx_multi_sheet_without_choice_raises(tmp_path):
    f = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a"])
    ws1.append(["1"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["b"])
    ws2.append(["2"])
    wb.save(f)

    with pytest.raises(IngestFormatError, match="First.*Second|Second.*First"):
        read_excel_rows(f)


def test_read_excel_rows_xlsx_multi_sheet_with_explicit_choice_succeeds(tmp_path):
    f = tmp_path / "multi2.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a"])
    ws1.append(["1"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["b"])
    ws2.append(["2"])
    wb.save(f)

    assert read_excel_rows(f, sheet="Second") == [{"b": "2"}]


def test_read_excel_rows_xlsx_unknown_sheet_name_raises(tmp_path):
    f = tmp_path / "single.xlsx"
    _write_xlsx(f, [["a"], ["1"]])
    with pytest.raises(IngestFormatError, match="no sheet named"):
        read_excel_rows(f, sheet="DoesNotExist")


def test_list_excel_sheets_single_sheet(tmp_path):
    f = tmp_path / "single.xlsx"
    _write_xlsx(f, [["a"], ["1"]], sheet_name="OnlySheet")
    assert list_excel_sheets(f) == ["OnlySheet"]


def test_read_excel_rows_xls_happy_path(tmp_path):
    f = tmp_path / "sample.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    for r, row in enumerate([["participant_id", "age"], ["001", 34], ["002", 29]]):
        for c, value in enumerate(row):
            ws.write(r, c, value)
    wb.save(str(f))

    rows = read_excel_rows(f)
    assert rows == [
        {"participant_id": "001", "age": "34"},
        {"participant_id": "002", "age": "29"},
    ]


def test_read_excel_rows_xls_skips_blank_formatted_trailing_row(tmp_path):
    # A real .xls file can carry a trailing row whose cells are XL_CELL_BLANK
    # (formatting like a border applied to an otherwise-empty cell), not
    # XL_CELL_EMPTY -- both must be treated as "no value" so the existing
    # fully-blank-row skip in read_excel_rows fires instead of producing a
    # row of empty strings.
    f = tmp_path / "blank_row.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "a")
    ws.write(0, 1, "b")
    ws.write(1, 0, "1")
    ws.write(1, 1, "2")
    blank_style = xlwt.easyxf("borders: top thin")
    ws.write(2, 0, "", blank_style)
    ws.write(2, 1, "", blank_style)
    wb.save(str(f))

    assert read_excel_rows(f) == [{"a": "1", "b": "2"}]


def test_read_excel_rows_xls_date_only_cell_has_no_time_suffix(tmp_path):
    f = tmp_path / "date_only.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "visit_date")
    ws.write(1, 0, dt.date(2024, 1, 15), xlwt.easyxf(num_format_str="YYYY-MM-DD"))
    wb.save(str(f))

    assert read_excel_rows(f) == [{"visit_date": "2024-01-15"}]
