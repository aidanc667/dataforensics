from pathlib import Path

import pytest

from dataforensics.dictionary import build_data_dictionary, count_stripped_footer_lines, read_rows
from dataforensics.ingest import DuplicateHeaderError


def test_dictionary_basic_fields(tmp_path):
    f = tmp_path / "sample.csv"
    f.write_text(
        "participant_id,sex,age,notes\n"
        "001,M,34,fine\n"
        "002,F,29,fine\n"
        "003,M,,fine\n"
        "004,F,41,fine\n"
    )
    d = build_data_dictionary(f)

    assert d["participant_id"]["category"] == "id"
    assert d["participant_id"]["dtype"] == "Utf8"

    assert d["age"]["null_count"] == 1
    assert d["age"]["non_null_pct"] == 75.0

    assert d["sex"]["category"] == "categorical"
    assert set(d["sex"]["levels"]) == {"M", "F"}


def test_dictionary_zero_variance_flag(tmp_path):
    f = tmp_path / "flat.csv"
    f.write_text("id,site\n001,A\n002,A\n003,A\n")
    d = build_data_dictionary(f)
    assert d["site"]["is_zero_variance"] is True
    assert d["id"]["is_zero_variance"] is False


def test_dictionary_zero_vs_null_kept_separate(tmp_path):
    f = tmp_path / "smoking.csv"
    f.write_text("id,cigs_per_day\n001,0\n002,\n003,5\n")
    d = build_data_dictionary(f)
    assert d["cigs_per_day"]["zero_count"] == 1
    assert d["cigs_per_day"]["null_count"] == 1


def test_dictionary_unique_count_ignores_leading_trailing_whitespace(tmp_path):
    # "Toronto" and "Toronto " (trailing space) are the same city, not two
    # different ones -- a handful of whitespace-padded rows should not
    # inflate unique_count past the real number of distinct values.
    f = tmp_path / "city.csv"
    f.write_text(
        "id,city\n"
        "1,Toronto\n"
        "2, Toronto\n"
        "3,Toronto \n"
        "4,Berlin\n"
        "5,Berlin\n"
    )
    d = build_data_dictionary(f)
    assert d["city"]["unique_count"] == 2
    assert set(d["city"]["levels"]) == {"Toronto", "Berlin"}


def test_dictionary_whitespace_padding_does_not_hide_levels_past_cardinality_cap(tmp_path):
    # Regression: a genuinely low-cardinality column (6 real cities across
    # 200 rows, well under cardinality_cap(200) == 10) was being
    # misclassified "free_text" (levels hidden) purely because whitespace
    # padding on some rows inflated the raw unique count past the cap.
    cities = ["Toronto", "Berlin", "Lisbon", "Nairobi", "Osaka", "Bangalore"]
    lines = ["id,city"]
    for i in range(200):
        city = cities[i % len(cities)]
        if i % 7 == 0:
            city = city + " "  # trailing-whitespace variant of a real city
        lines.append(f"{i + 1},{city}")
    f = tmp_path / "cities200.csv"
    f.write_text("\n".join(lines) + "\n")
    d = build_data_dictionary(f)
    assert d["city"]["category"] == "categorical"
    assert set(d["city"]["levels"]) == set(cities)


def test_dictionary_zero_count_ignores_whitespace_padding(tmp_path):
    f = tmp_path / "smoking.csv"
    f.write_text("id,cigs_per_day\n001, 0\n002,0 \n003,5\n")
    d = build_data_dictionary(f)
    assert d["cigs_per_day"]["zero_count"] == 2


def test_dictionary_skips_top_code_detection_on_zero_variance_categorical_column(tmp_path):
    # Regression: a boolean-flag-shaped column (every non-null value
    # literally "1") got classified "categorical" (unique_count == 1) but
    # STILL ran numeric top-code detection, which trivially fires 100%
    # of the time on any single-value column -- "100% of values sit at
    # the observed maximum" is meaningless when there's only one value to
    # begin with; that's already captured by is_zero_variance.
    lines = ["id,flag"] + [f"{i},1" for i in range(1, 21)]
    f = tmp_path / "flag.csv"
    f.write_text("\n".join(lines) + "\n")
    d = build_data_dictionary(f)
    assert d["flag"]["category"] == "categorical"
    assert d["flag"]["is_zero_variance"] is True
    assert d["flag"]["top_code_spike"] is None
    assert d["flag"]["outliers"] is None


def test_dictionary_skips_outlier_detection_on_low_cardinality_numeric_coded_column(tmp_path):
    # Regression: a low-cardinality numeric-coded column (e.g. a district
    # number 1-11) is a small closed set of discrete labels, not a
    # continuum -- IQR-based outlier/top-code detection doesn't apply to
    # it any more than it would to a column of city names. Statistical
    # detection should only run for "free_text" (high-enough-cardinality
    # to look continuous) numeric columns.
    lines = ["id,district"] + [f"{i},{(i % 11) + 1}" for i in range(1, 301)]
    f = tmp_path / "district.csv"
    f.write_text("\n".join(lines) + "\n")
    d = build_data_dictionary(f)
    assert d["district"]["category"] == "categorical"
    assert d["district"]["top_code_spike"] is None
    assert d["district"]["outliers"] is None


def test_dictionary_still_runs_top_code_detection_on_high_cardinality_numeric_column(tmp_path):
    # Confirms the fix didn't overreach: a genuinely high-cardinality
    # numeric column (classified "free_text") must still get outlier/
    # top-code detection -- that's the exact case this feature exists for.
    lines = ["id,income"] + [f"{i},{30000 + i}" for i in range(1, 20)] + ["20,250000", "21,250000", "22,250000"]
    f = tmp_path / "income.csv"
    f.write_text("\n".join(lines) + "\n")
    d = build_data_dictionary(f)
    assert d["income"]["category"] == "free_text"
    assert d["income"]["top_code_spike"] is not None
    assert d["income"]["top_code_spike"]["value"] == 250000.0


def test_dictionary_high_cardinality_is_free_text(tmp_path):
    rows = "\n".join(f"{i},note-{i}-unique" for i in range(60))
    f = tmp_path / "notes.csv"
    f.write_text("id,note\n" + rows + "\n")
    d = build_data_dictionary(f)
    assert d["note"]["category"] == "free_text"


def test_dictionary_small_sample_categorical_not_suppressed_by_ratio_cap(tmp_path):
    # With only 4 data rows, a naive 5%-of-N cardinality cap rounds to 0/1,
    # which would wrongly push a normal 2-level column like "sex" into
    # free_text. A floor on the cap keeps small files usable.
    f = tmp_path / "sample.csv"
    f.write_text(
        "id,sex\n001,M\n002,F\n003,M\n004,F\n"
    )
    d = build_data_dictionary(f)
    assert d["sex"]["category"] == "categorical"
    assert set(d["sex"]["levels"]) == {"M", "F"}


def test_dictionary_ragged_row_counts_missing_trailing_field_as_null(tmp_path):
    # A data row with fewer fields than the header (a common malformed-export
    # pattern) must count the missing trailing field as null rather than
    # silently dropping it from that column's value list.
    f = tmp_path / "ragged.csv"
    f.write_text("a,b,c\n1,2,3\n4,5\n")
    d = build_data_dictionary(f)
    assert d["c"]["null_count"] == 1
    assert d["c"]["non_null_pct"] == 50.0


def test_read_rows_ragged_row_fills_missing_trailing_field(tmp_path):
    f = tmp_path / "ragged.csv"
    f.write_text("participant_id,age,site\n1,40,A\n2,41\n")  # row 2 is missing the 'site' field
    rows = read_rows(f)
    assert rows[1] == {"participant_id": "2", "age": "41", "site": ""}


def test_read_rows_row_with_one_overflow_field_preserves_it(tmp_path):
    f = tmp_path / "overflow.csv"
    f.write_text("id,name\n1,Ada,stray\n")  # an unquoted stray delimiter -- 3 raw fields against a 2-field header
    rows = read_rows(f)
    assert rows[0] == {"id": "1", "name": "Ada", "": "stray"}


def test_read_rows_row_with_multiple_overflow_fields_preserves_all_of_them(tmp_path):
    # Regression test for a real, serious bug: dict(zip_longest(header, row,
    # fillvalue="")) pairs EVERY position past the header's length with the
    # SAME key ("", since fillvalue is reused once header is exhausted) --
    # dict() construction then keeps only the LAST such pair, silently
    # discarding every earlier overflow value with zero indication anything
    # vanished. A free-text field with two unquoted/unescaped delimiters
    # (e.g. "Springfield, IL, USA" in an unquoted comma-delimited row) would
    # have lost "Springfield, IL" and kept only "USA". Both overflow
    # fragments must now survive, joined together rather than dropped.
    f = tmp_path / "multi_overflow.csv"
    f.write_text("id,name\n1,Ada,overflow1,overflow2\n")
    rows = read_rows(f)
    assert rows[0]["id"] == "1"
    assert rows[0]["name"] == "Ada"
    assert "overflow1" in rows[0][""]
    assert "overflow2" in rows[0][""]


def test_read_rows_empty_file_returns_empty_list(tmp_path):
    f = tmp_path / "empty.csv"
    f.write_text("")
    assert read_rows(f) == []


def test_read_rows_duplicate_header_raises_instead_of_silently_dropping_data(tmp_path):
    # header "pid,sex,sex" with values "1,M,F": without duplicate-header
    # detection, dict(zip_longest(...)) collapses both "sex" columns into
    # one key and the "M" value is silently lost with no error, exit 0.
    f = tmp_path / "dup.csv"
    f.write_text("pid,sex,sex\n1,M,F\n")
    with pytest.raises(DuplicateHeaderError, match="sex"):
        read_rows(f)


def test_build_data_dictionary_duplicate_header_raises(tmp_path):
    f = tmp_path / "dup.csv"
    f.write_text("pid,sex,sex\n1,M,F\n")
    with pytest.raises(DuplicateHeaderError, match="sex"):
        build_data_dictionary(f)


import json


def test_build_data_dictionary_from_json(tmp_path):
    f = tmp_path / "sample.json"
    f.write_text(json.dumps([
        {"participant_id": "001", "sex": "M", "age": 34, "notes": "fine"},
        {"participant_id": "002", "sex": "F", "age": 29, "notes": "fine"},
        {"participant_id": "003", "sex": "M", "age": None, "notes": "fine"},
        {"participant_id": "004", "sex": "F", "age": 41, "notes": "fine"},
    ]))
    d = build_data_dictionary(f)

    assert d["participant_id"]["category"] == "id"
    assert d["age"]["null_count"] == 1
    assert d["age"]["non_null_pct"] == 75.0
    assert d["sex"]["category"] == "categorical"
    assert set(d["sex"]["levels"]) == {"M", "F"}


def test_read_rows_from_json(tmp_path):
    f = tmp_path / "sample.json"
    f.write_text(json.dumps([{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]))
    assert read_rows(f) == [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]


def test_build_data_dictionary_from_excel(tmp_path):
    import openpyxl

    f = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["participant_id", "sex", "age"])
    ws.append(["001", "M", 34])
    ws.append(["002", "F", 29])
    ws.append(["003", "M", None])
    ws.append(["004", "F", 41])
    wb.save(f)

    d = build_data_dictionary(f)
    assert d["participant_id"]["category"] == "id"
    assert d["age"]["null_count"] == 1
    assert d["sex"]["category"] == "categorical"


def test_read_rows_from_excel_with_explicit_sheet(tmp_path):
    import openpyxl

    f = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a"])
    ws1.append(["1"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["b"])
    ws2.append(["2"])
    wb.save(f)

    assert read_rows(f, sheet="Second") == [{"b": "2"}]


def test_build_data_dictionary_from_header_only_excel_preserves_schema(tmp_path):
    # A header-only Excel sheet (e.g. a blank data-collection template) has a
    # real, physically-present header row with zero data rows below it --
    # unlike a genuinely empty sheet, which has no header at all. The data
    # dictionary should reflect that real schema with zero-row stats, not
    # collapse to {} as if the file were totally empty.
    import openpyxl

    f = tmp_path / "header_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["a", "b", "c"])
    wb.save(f)

    d = build_data_dictionary(f)
    assert set(d.keys()) == {"a", "b", "c"}
    assert d["a"]["null_count"] == 0
    assert d["a"]["non_null_pct"] == 0.0


def test_build_data_dictionary_empty_json_array_returns_empty_dict(tmp_path):
    f = tmp_path / "empty.json"
    f.write_text("[]")
    assert build_data_dictionary(f) == {}


def test_find_outlier_evidence_returns_real_row_indices():
    from dataforensics.dictionary import find_outlier_evidence

    # A tight cluster around 30-34 plus one genuine outlier at row index 5.
    rows = [
        {"age": "30"}, {"age": "31"}, {"age": "32"}, {"age": "33"}, {"age": "34"},
        {"age": "300"},
    ]
    evidence = find_outlier_evidence(rows, "age")
    assert evidence == [(5, "300")]


def test_find_outlier_evidence_empty_for_non_numeric_column():
    from dataforensics.dictionary import find_outlier_evidence

    rows = [{"name": "Bob"}, {"name": "not-a-number"}]
    assert find_outlier_evidence(rows, "name") == []


def test_find_outlier_evidence_skips_nulls():
    from dataforensics.dictionary import find_outlier_evidence

    rows = [{"age": "30"}, {"age": ""}, {"age": "31"}, {"age": "32"}, {"age": "300"}]
    evidence = find_outlier_evidence(rows, "age")
    # row index 4 (not 3) is the outlier, since the blank at index 1 is
    # correctly skipped rather than shifting every later index down by one.
    assert evidence == [(4, "300")]


def test_find_outlier_evidence_empty_when_column_contains_nan_string():
    from dataforensics.dictionary import find_outlier_evidence

    # Regression test: Python's bare float() accepts the literal string
    # "nan" without raising, so a naive "all values must parse" check
    # would have treated this column as uniformly numeric and let NaN
    # corrupt detect_outliers' sort-based min/max/median/quartiles. Must
    # be treated the same as any other non-numeric value: not numeric.
    rows = [{"age": "30"}, {"age": "nan"}, {"age": "31"}, {"age": "300"}]
    assert find_outlier_evidence(rows, "age") == []


def test_build_data_dictionary_column_with_nan_string_is_not_treated_as_numeric(tmp_path):
    f = tmp_path / "sample.csv"
    f.write_text("id,age\n1,30\n2,nan\n3,45\n4,inf\n5,50\n")
    d = build_data_dictionary(f)
    # Falls back to categorical (low cardinality), never "numeric" with a
    # NaN/inf silently baked into its outlier/top-code statistics.
    assert d["age"]["outliers"] is None
    assert d["age"]["top_code_spike"] is None


def test_read_rows_does_not_drop_data_after_a_mid_file_blank_and_ragged_row(tmp_path):
    # Regression test for a real user-reported bug: id,name.csv with 5 real
    # rows, a blank line, then one ragged (2-field, missing "name") row,
    # then more real rows -- previously silently truncated everything from
    # the ragged row onward because it looked like the start of a footer.
    f = tmp_path / "sample.csv"
    f.write_text(
        "id,name,city\n"
        "1,Ada,Berlin\n"
        "2,Bob,Osaka\n"
        "\n"
        "3,Carol\n"
        "4,Dan,Lisbon\n"
        "5,Eve,Toronto\n"
    )
    rows = read_rows(f)
    assert len(rows) == 5
    assert rows[-1] == {"id": "5", "name": "Eve", "city": "Toronto"}
    assert rows[2] == {"id": "3", "name": "Carol", "city": ""}


def test_count_stripped_footer_lines_zero_when_nothing_stripped(tmp_path):
    f = tmp_path / "clean.csv"
    f.write_text("id,name\n1,Ada\n2,Bob\n")
    assert count_stripped_footer_lines(f) == 0


def test_count_stripped_footer_lines_counts_a_genuine_footer(tmp_path):
    f = tmp_path / "with_footer.csv"
    f.write_text("id,name\n1,Ada\n2,Bob\nQuery Parameters:\nGroup By: id\n")
    assert count_stripped_footer_lines(f) == 2


def test_count_stripped_footer_lines_zero_for_mid_file_noise_not_a_real_footer(tmp_path):
    # A blank line followed by one ragged row, with real data resuming
    # after -- must NOT be misread as a footer (the exact bug this
    # function's caller warns about must not fire on ordinary ragged data).
    f = tmp_path / "ragged.csv"
    f.write_text("id,name\n1,Ada\n2,Bob\n\n3\n4,Carol\n5,Dan\n")
    assert count_stripped_footer_lines(f) == 0


def test_find_top_code_evidence_returns_rows_at_the_ceiling():
    from dataforensics.dictionary import find_top_code_evidence

    rows = [{"age": "34"}, {"age": "99"}, {"age": "40"}, {"age": "99"}]
    evidence = find_top_code_evidence(rows, "age", 99.0)
    assert evidence == [(1, "99"), (3, "99")]
