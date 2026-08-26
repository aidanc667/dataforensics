import json
from pathlib import Path

from click.testing import CliRunner

from dataforensics.cli import main
from dataforensics.hashing import sha256_file


def test_scan_writes_dictionary_and_never_modifies_input(tmp_path):
    src = tmp_path / "sample.csv"
    src.write_text("id,age\n001,34\n002,29\n")
    before_hash = sha256_file(src)

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])

    assert result.exit_code == 0
    assert sha256_file(src) == before_hash

    json_path = tmp_path / "sample.data_dictionary.json"
    md_path = tmp_path / "sample.data_dictionary.md"
    assert json_path.exists()
    assert md_path.exists()

    payload = json.loads(json_path.read_text())
    assert "age" in payload


def test_scan_warns_on_stderr_when_footer_stripping_drops_lines(tmp_path):
    # strip_footer's field-count heuristic is not CSV-quote-aware (see
    # ingest.strip_footer's docs): two consecutive genuine data rows
    # containing a quoted delimiter (a comma inside quotes, in a
    # comma-delimited file) raise the raw comma count above the header's,
    # so strip_footer misclassifies both as a footer block and drops them.
    # This is a known, documented limitation (see README's "Known
    # limitations") -- the fix under test isn't preventing the
    # misclassification, only making sure it's no longer silent.
    src = tmp_path / "clinics.csv"
    src.write_text(
        "name,age\n"
        "Bob,34\n"
        '"Delta Clinic, North",40\n'
        '"Acme Labs, South",50\n'
    )

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])

    assert result.exit_code == 0
    assert "Warning" in result.output
    assert "2" in result.output  # 2 lines stripped
    assert "clinics.csv" in result.output

    # The misclassification itself is real: both genuine data rows are gone
    # from the dictionary's row-derived stats, confirming this reproduces
    # the documented limitation rather than testing a scenario that can't
    # happen.
    payload = json.loads((tmp_path / "clinics.data_dictionary.json").read_text())
    assert payload["name"]["null_count"] == 0
    assert payload["age"]["unique_count"] == 1  # only "34" (Bob) survived


def test_scan_prints_no_warning_when_nothing_stripped(tmp_path):
    src = tmp_path / "clean.csv"
    src.write_text("id,age\n001,34\n002,29\n")

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])

    assert result.exit_code == 0
    assert "Warning" not in result.output


import openpyxl


def test_scan_accepts_json_input(tmp_path):
    src = tmp_path / "sample.json"
    src.write_text(json.dumps([{"id": "001", "age": 34}, {"id": "002", "age": 29}]))

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])

    assert result.exit_code == 0
    payload = json.loads((tmp_path / "sample.data_dictionary.json").read_text())
    assert "age" in payload


def test_scan_malformed_json_exits_3(tmp_path):
    src = tmp_path / "broken.json"
    src.write_text("{not valid")

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])

    assert result.exit_code == 3
    assert "Malformed input file" in result.output


def test_scan_accepts_xlsx_input(tmp_path):
    src = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "age"])
    ws.append(["001", 34])
    ws.append(["002", 29])
    wb.save(src)

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])

    assert result.exit_code == 0
    payload = json.loads((tmp_path / "sample.data_dictionary.json").read_text())
    assert "age" in payload


def test_scan_xlsx_multi_sheet_requires_sheet_option(tmp_path):
    src = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["a"])
    ws1.append(["1"])
    ws2 = wb.create_sheet("Second")
    ws2.append(["b"])
    ws2.append(["2"])
    wb.save(src)

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path)])
    assert result.exit_code == 3
    assert "multiple sheets" in result.output

    result = CliRunner().invoke(main, ["scan", str(src), "--out-dir", str(tmp_path), "--sheet", "Second"])
    assert result.exit_code == 0
    payload = json.loads((tmp_path / "multi.data_dictionary.json").read_text())
    assert "b" in payload


def test_harmonize_execute_on_header_only_xlsx_preserves_header(tmp_path):
    # Regression test for the final-review finding: cli.py's
    # _read_header_and_row_count used to call read_excel_rows for its
    # excel branch, which returns list[dict[str, str]] -- a shape that
    # can't distinguish "no header at all" from "header present, zero
    # data rows" (both come back as []). On a header-only .xlsx (a real
    # header row, zero data rows), that collapsed the anchor's header to
    # [], which fed into _harmonize_single_file's `fieldnames =
    # anchor_header` fallback used when transformed_rows is empty --
    # writing an output file with no header line at all, silently, at
    # exit 0. It must instead use read_excel_table (like
    # dictionary.py's _load_table already does) so the header survives.
    src = tmp_path / "header_only.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["participant_id", "smoking_status"])
    wb.save(src)

    rules_path = tmp_path / "rules.yaml"
    rules_path.write_text("version: 1\nprimary_key: [participant_id]\ncolumns: {}\n")
    output_path = tmp_path / "out.csv"

    result = CliRunner().invoke(
        main,
        [
            "harmonize",
            str(src),
            "--rules",
            str(rules_path),
            "--output",
            str(output_path),
            "--execute",
        ],
    )

    assert result.exit_code == 0, result.output
    assert output_path.exists()
    first_line = output_path.read_text().splitlines()[0]
    assert first_line == "participant_id,smoking_status"
